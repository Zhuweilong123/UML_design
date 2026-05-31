# -*- coding: utf-8 -*-
"""Generate test case Excel from UML design file."""
import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Read UML file ───────────────────────────────────
uml_path = 'D:/AI_tools/uml_designer/backend/uml_files/uml_design_3_opt.uml'
with open(uml_path, 'r', encoding='utf-8') as f:
    uml = json.load(f)

print(f"Loaded UML: {len(uml['classes'])} classes, {len(uml['relations'])} relations")

# ── Style definitions ───────────────────────────────
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
normal_font = Font(name='Microsoft YaHei', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
headers = ['用例ID', '用例名称', '优先级', '前置条件', '测试步骤', '预期结果', '测试类型', '状态']

def style_sheet(ws, col_widths, hdrs):
    for i, (w, h) in enumerate(zip(col_widths, hdrs), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

def write_rows(ws, rows, start_row=2):
    for r, row in enumerate(rows, start_row):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = wrap_align

wb = Workbook()

# ═══════════════════════════════════════════════════════
# Sheet 1: Overview
# ═══════════════════════════════════════════════════════
ws = wb.active
ws.title = '用例概览'
style_sheet(ws, [15, 15, 10, 10, 18, 25, 30], ['模块', '类名', '构造型', '用例数', '优先级分布', '关系', '备注'])
summary = [
    ['任务基类', 'BaseTask', 'abstract', '7', 'P0:3, P1:2, P2:2', 'OtaTask/CrowTask/SentinelTask继承', '抽象基类，execute()为抽象方法'],
    ['升级任务', 'OtaTask', 'class', '8', 'P0:4, P1:3, P2:1', '继承BaseTask, 关联TaskScheduler', '随机升级，清除鸡叫累计和预约'],
    ['鸡叫任务', 'CrowTask', 'class', '10', 'P0:5, P1:3, P2:2', '继承BaseTask, 聚合到TaskScheduler', '每7天鸡叫，2:00-4:00随机，不可打断'],
    ['哨兵任务', 'SentinelTask', 'class', '9', 'P0:4, P1:3, P2:2', '继承BaseTask, 关联TaskScheduler', '可被CrowTask/OtaTask打断并恢复'],
    ['任务调度器', 'TaskScheduler', 'class', '11', 'P0:5, P1:4, P2:2', '聚合到MM_APP, 聚合CrowTask', '任务增删调度，按优先级执行'],
    ['主应用', 'MM_APP', 'class', '6', 'P0:3, P1:2, P2:1', '聚合TaskScheduler', '应用启停入口'],
    ['集成测试', '-', '-', '12', 'P0:6, P1:4, P2:2', '全链路', '跨模块集成场景'],
]
write_rows(ws, summary)

# ═══════════════════════════════════════════════════════
# Sheet 2: BaseTask Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('BaseTask')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-BASE-001', '创建BaseTask子类实例', 'P0', '无', '1.创建OtaTask实例\n2.检查taskId不为空\n3.检查priority默认值为0', '实例创建成功，taskId自动生成，priority=0', '单元测试', '待执行'],
    ['TC-BASE-002', 'getTaskId返回正确ID', 'P0', 'BaseTask子类已创建', '1.创建CrowTask实例\n2.调用getTaskId()\n3.验证返回值', '返回的taskId与创建时一致', '单元测试', '待执行'],
    ['TC-BASE-003', '子类重写execute方法', 'P0', '各子类已创建', '1.调用OtaTask.execute()\n2.调用CrowTask.execute()\n3.调用SentinelTask.execute()', '各子类执行各自execute逻辑，不抛异常', '单元测试', '待执行'],
    ['TC-BASE-004', 'BaseTask不可直接实例化', 'P1', '无', '1.尝试直接创建BaseTask实例\n2.调用execute()', '创建失败或execute()抛NotImplementedError', '单元测试', '待执行'],
    ['TC-BASE-005', 'priority边界值测试', 'P1', '子类已创建', '1.设置priority=Integer.MAX\n2.设置priority=Integer.MIN\n3.设置priority=-1', '边界值正常存储，不报错', '单元测试', '待执行'],
    ['TC-BASE-006', '多子类taskId唯一性', 'P2', '无', '1.创建OtaTask\n2.创建CrowTask\n3.创建SentinelTask\n4.比较三者taskId', '三个taskId互不相同', '单元测试', '待执行'],
    ['TC-BASE-007', 'protected属性子类访问', 'P2', '子类已创建', '1.在CrowTask中访问priority\n2.修改priority值', '子类可读写protected属性', '单元测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 3: OtaTask Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('OtaTask')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-OTA-001', 'OtaTask.execute正常执行', 'P0', 'OtaTask实例已创建', '1.调用OtaTask.execute()\n2.检查升级流程启动', '升级流程正常启动，参数正确', '单元测试', '待执行'],
    ['TC-OTA-002', 'isRandom配置生效', 'P0', 'OtaTask.isRandom已设置', '1.isRandom=true多次执行\n2.收集升级时间点\n3.isRandom=false再测试', 'true时升级时间随机；false时固定', '单元测试', '待执行'],
    ['TC-OTA-003', '升级后清除鸡叫累计', 'P0', '鸡叫累计>0', '1.设置鸡叫累计值>0\n2.执行OtaTask\n3.检查鸡叫累计值', '升级后鸡叫累计值归零', '集成测试', '待执行'],
    ['TC-OTA-004', '升级后清除预约鸡叫', 'P0', '有预约的鸡叫', '1.预约一个鸡叫请求\n2.执行OtaTask\n3.检查预约队列', '升级后预约鸡叫队列清空', '集成测试', '待执行'],
    ['TC-OTA-005', '升级中CrowTask到达', 'P1', 'OtaTask正在执行', '1.OtaTask启动\n2.CrowTask到达预定时间\n3.检查执行顺序', '根据优先级决定是否打断', '集成测试', '待执行'],
    ['TC-OTA-006', '设备离线时升级', 'P1', '设备离线', '1.模拟设备离线\n2.执行OtaTask', '升级失败并记录离线日志', '单元测试', '待执行'],
    ['TC-OTA-007', '重复升级跳过', 'P1', 'OtaTask已执行过', '1.再次执行OtaTask\n2.检查是否重复', '已有最新版本时跳过升级', '单元测试', '待执行'],
    ['TC-OTA-008', '升级中系统重启', 'P2', 'OtaTask执行到一半', '1.升级50%时重启\n2.重启后检查状态', '重启后恢复或重新开始', '集成测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 4: CrowTask Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('CrowTask')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-CROW-001', 'scheduleNextCrow正常调度', 'P0', 'CrowTask实例已创建', '1.设置intervalDays=7\n2.scheduleNextCrow()\n3.检查下次时间', '下次鸡叫=当前+7天，2:00-4:00间', '单元测试', '待执行'],
    ['TC-CROW-002', '鸡叫时间窗口2:00-4:00', 'P0', 'CrowTask已调度', '1.到达间隔时间触发\n2.多次检查实际时间', '每次在2:00-4:00范围内随机', '单元测试', '待执行'],
    ['TC-CROW-003', '鸡叫不可打断验证', 'P0', 'CrowTask正在鸡叫', '1.鸡叫中\n2.SentinelTask尝试打断\n3.OtaTask尝试打断', '鸡叫不接受打断请求', '集成测试', '待执行'],
    ['TC-CROW-004', 'clearCrowFlag清除标志', 'P0', 'crowFlag=True', '1.调用clearCrowFlag()\n2.检查crowFlag', 'crowFlag变为false', '单元测试', '待执行'],
    ['TC-CROW-005', '间隔天数边界值', 'P0', 'CrowTask已创建', '1.intervalDays=0\n2.intervalDays=1\n3.intervalDays=365', '0立即执行；1和365正常', '单元测试', '待执行'],
    ['TC-CROW-006', 'scheduledTime受保护', 'P1', 'CrowTask已创建', '1.外部代码访问scheduledTime\n2.子类代码访问', '外部不可访问；子类可访问', '单元测试', '待执行'],
    ['TC-CROW-007', '鸡叫与升级同时', 'P1', '鸡叫+升级同时触发', '1.同时触发\n2.检查执行顺序', '鸡叫优先执行', '集成测试', '待执行'],
    ['TC-CROW-008', '预约鸡叫被清除', 'P1', '有预约的鸡叫', '1.调用clearScheduledCrow\n2.等待原鸡叫时间', '鸡叫不再执行', '单元测试', '待执行'],
    ['TC-CROW-009', '鸡叫累计值管理', 'P2', '系统有鸡叫累计', '1.多次触发鸡叫\n2.检查累计值\n3.清除累计', '累计正确递增，可清除', '单元测试', '待执行'],
    ['TC-CROW-010', '跨天调度', 'P2', '间隔7天', '1.模拟跨23:59:59\n2.检查调度', '跨天后时间计算正确', '单元测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 5: SentinelTask Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('SentinelTask')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-SEN-001', '哨兵正常执行', 'P0', 'SentinelTask已创建', '1.调用execute()\n2.检查isRunning', 'isRunning=true', '单元测试', '待执行'],
    ['TC-SEN-002', '哨兵被CrowTask打断', 'P0', '哨兵正在执行', '1.SentinelTask.execute()\n2.CrowTask触发\n3.检查isRunning+interruptedBy', 'isRunning=false, interruptedBy=CrowTask', '集成测试', '待执行'],
    ['TC-SEN-003', '哨兵被OtaTask打断', 'P0', '哨兵正在执行', '1.SentinelTask.execute()\n2.OtaTask触发\n3.检查打断状态', '哨兵被打断，状态正确保存', '集成测试', '待执行'],
    ['TC-SEN-004', '哨兵恢复执行resume()', 'P0', '哨兵被打断后', '1.打断任务完成\n2.调用resume()\n3.检查isRunning', 'isRunning恢复true，之前状态恢复', '单元测试', '待执行'],
    ['TC-SEN-005', '未运行时打断信号', 'P1', '哨兵未启动', '1.调用interrupt()\n2.检查状态', '打断被忽略，isRunning仍为false', '单元测试', '待执行'],
    ['TC-SEN-006', '多次打断+恢复', 'P1', '哨兵反复被打断', '1.打断->恢复 重复3次\n2.检查状态', '每次恢复后状态正确，无残留', '单元测试', '待执行'],
    ['TC-SEN-007', '优先级打断规则', 'P1', '多种任务并存', '1.高/同/低优任务到达\n2.检查是否打断', '高优打断；同优FIFO；低优等待', '集成测试', '待执行'],
    ['TC-SEN-008', '打断异常后恢复', 'P2', '哨兵被打断', '1.打断任务抛异常\n2.检查哨兵\n3.resume()', '异常不影响哨兵状态恢复', '集成测试', '待执行'],
    ['TC-SEN-009', 'interruptedBy封装', 'P2', '哨兵被打断', '1.外部直接修改interruptedBy\n2.检查', '外部不可直接修改(private)', '单元测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 6: TaskScheduler Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('TaskScheduler')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-SCH-001', 'addTask添加任务', 'P0', 'TaskScheduler初始化', '1.addTask(OtaTask)\n2.addTask(CrowTask)\n3.检查taskList', 'taskList长度=2', '单元测试', '待执行'],
    ['TC-SCH-002', 'removeTask删除任务', 'P0', 'taskList有2个任务', '1.removeTask(OtaTask.id)\n2.检查taskList', 'OtaTask被移除', '单元测试', '待执行'],
    ['TC-SCH-003', '按优先级排序执行', 'P0', '不同优先级任务各1', '1.添加P0/P1/P2任务\n2.executeTasks()\n3.记录顺序', 'P0->P1->P2顺序执行', '单元测试', '待执行'],
    ['TC-SCH-004', '空列表执行', 'P0', 'taskList为空', '1.调用executeTasks()', '不抛异常，正常返回', '单元测试', '待执行'],
    ['TC-SCH-005', '执行中动态添加', 'P0', 'executeTasks()执行中', '1.执行中添加新任务\n2.检查是否执行', '当前轮不执行新任务', '集成测试', '待执行'],
    ['TC-SCH-006', '删除不存在任务', 'P1', 'taskList有2任务', '1.removeTask(不存在ID)', 'taskList不变或抛指定异常', '单元测试', '待执行'],
    ['TC-SCH-007', '重复添加去重', 'P1', 'taskList已有OtaTask', '1.再次addTask(同OtaTask)', '不可重复添加或自动去重', '单元测试', '待执行'],
    ['TC-SCH-008', '同优先级FIFO', 'P1', '3个P1任务', '1.A->B->C顺序添加\n2.执行', 'A->B->C顺序执行', '单元测试', '待执行'],
    ['TC-SCH-009', '任务执行超时', 'P1', '某任务执行长', '1.添加长任务\n2.executeTasks加超时', '超时后标记失败继续下一个', '集成测试', '待执行'],
    ['TC-SCH-010', '全部完成回调', 'P2', '3个任务', '1.注册回调\n2.executeTasks()\n3.验证', '全部完成后触发回调', '单元测试', '待执行'],
    ['TC-SCH-011', '并发addTask安全', 'P2', '多线程', '1.3线程同时addTask\n2.检查完整性', '无竞态，所有任务成功添加', '单元测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 7: MM_APP Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('MM_APP')
style_sheet(ws, [12, 24, 8, 20, 38, 30, 12, 10], headers)
cases = [
    ['TC-APP-001', '应用正常启动', 'P0', 'MM_APP已创建', '1.start()\n2.检查调度器初始化\n3.检查任务注册', '调度器初始化，任务注册完毕', '集成测试', '待执行'],
    ['TC-APP-002', '应用正常停止', 'P0', 'MM_APP运行中', '1.stop()\n2.检查任务停止\n3.检查资源释放', '所有任务停止，资源释放', '集成测试', '待执行'],
    ['TC-APP-003', '获取应用信息', 'P0', 'MM_APP已启动', '1.检查appName\n2.检查version', 'appName/version非空，格式正确', '单元测试', '待执行'],
    ['TC-APP-004', '重复启动防护', 'P1', 'MM_APP已启动', '1.再次start()', '不重复初始化', '单元测试', '待执行'],
    ['TC-APP-005', '未启动就停止', 'P1', 'MM_APP未启动', '1.直接stop()', '不抛异常', '单元测试', '待执行'],
    ['TC-APP-006', '启动失败回滚', 'P2', '依赖不可用', '1.模拟初始化失败\n2.start()\n3.检查状态', '启动失败时状态仍为stopped', '单元测试', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 8: Integration Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('集成测试')
style_sheet(ws, [12, 28, 8, 24, 42, 32, 12, 10], headers)
cases = [
    ['TC-INT-001', '完整启动-调度-停止流程', 'P0', '系统初始', '1.start()\n2.添加OtaTask+CrowTask\n3.executeTasks()\n4.stop()', '全流程无异常', '端到端', '待执行'],
    ['TC-INT-002', 'CrowTask打断SentinelTask', 'P0', '哨兵运行+鸡叫预约', '1.哨兵运行\n2.鸡叫到达\n3.被打断->鸡叫->恢复', '打断链: 暂停->鸡叫->恢复', '端到端', '待执行'],
    ['TC-INT-003', 'Ota升级清除鸡叫', 'P0', '鸡叫有累计', '1.触发OtaTask\n2.检查累计清零\n3.检查预约清除', '升级后鸡叫相关清零', '端到端', '待执行'],
    ['TC-INT-004', '鸡叫中升级到达', 'P0', 'CrowTask鸡叫中', '1.鸡叫中\n2.OtaTask升级触发\n3.检查', '鸡叫不可打断，升级等待', '端到端', '待执行'],
    ['TC-INT-005', '多任务并发调度', 'P0', '5个不同优先级任务', '1.同时到达执行时间\n2.调度\n3.检查顺序', '全部按优先级+到达顺序执行', '端到端', '待执行'],
    ['TC-INT-006', '7天稳定性测试', 'P0', '系统正常运行', '1.模拟运行7天\n2.每日鸡叫+随机升级\n3.检查内存', '无内存泄漏，无任务累积', '性能测试', '待执行'],
    ['TC-INT-007', '鸡叫累计循环', 'P1', '正常运行', '1.3次鸡叫->升级清除->2次鸡叫\n2.验证', '清除->再累计循环正确', '端到端', '待执行'],
    ['TC-INT-008', '三任务优先级仲裁', 'P1', '哨兵/鸡叫/升级同时', '1.哨兵执行中\n2.鸡叫+升级同时到达', '鸡叫(P0打断)->升级(P1)->哨兵恢复', '端到端', '待执行'],
    ['TC-INT-009', '异常任务后恢复', 'P1', '某任务会异常', '1.添加异常任务\n2.executeTasks()\n3.检查后续', '异常任务失败后续正常执行', '端到端', '待执行'],
    ['TC-INT-010', '启停压力测试', 'P1', '系统初始', '1.start/stop重复100次\n2.检查资源', '无泄漏', '性能测试', '待执行'],
    ['TC-INT-011', '时间跳变处理', 'P2', '运行中', '1.时间向前跳24h\n2.检查调度', '调度自动纠正', '端到端', '待执行'],
    ['TC-INT-012', '断电恢复', 'P2', '运行中', '1.任务执行中\n2.模拟断电\n3.重启检查', '重启后未完成恢复或标记失败', '端到端', '待执行'],
]
write_rows(ws, cases)

# ═══════════════════════════════════════════════════════
# Sheet 9: Relation Tests
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('关系验证')
style_sheet(ws, [12, 14, 14, 14, 36, 10], ['用例ID', '源类', '目标类', '关系类型', '验证方法', '状态'])

rel_cases = [
    ['TC-REL-001', 'OtaTask', 'BaseTask', 'inheritance', 'OtaTask instanceof BaseTask; 继承taskId/priority/execute()', '待验证'],
    ['TC-REL-002', 'CrowTask', 'BaseTask', 'inheritance', 'CrowTask instanceof BaseTask; 可调用getTaskId()', '待验证'],
    ['TC-REL-003', 'SentinelTask', 'BaseTask', 'inheritance', 'SentinelTask instanceof BaseTask; 可执行interrupt/resume', '待验证'],
    ['TC-REL-004', 'CrowTask', 'TaskScheduler', 'aggregation', 'TaskScheduler删除后CrowTask仍可独立存在', '待验证'],
    ['TC-REL-005', 'TaskScheduler', 'MM_APP', 'aggregation', 'MM_APP停止不影响调度器taskList数据', '待验证'],
    ['TC-REL-006', 'OtaTask', 'TaskScheduler', 'association', 'OtaTask注册到调度器后可被调度执行', '待验证'],
    ['TC-REL-007', 'SentinelTask', 'TaskScheduler', 'association', '哨兵通过调度器与其他任务协调', '待验证'],
]
write_rows(ws, rel_cases)

# ═══════════════════════════════════════════════════════
# Sheet 10: UML Design Info
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet('设计信息')
style_sheet(ws, [18, 18, 22, 30], ['类名', '构造型', '关键属性', '关键方法'])
for cls in uml['classes']:
    attrs = '\n'.join(f"{a['visibility']} {a['name']}: {a['type']}" for a in cls['attributes'])
    methods = '\n'.join(f"{m['visibility']} {m['name']}({m['params']}): {m['return_type']}" for m in cls['methods'])
    write_rows(ws, [[cls['name'], cls['stereotype'], attrs, methods]], start_row=ws.max_row+1)

# ── Save ────────────────────────────────────────────
outpath = 'D:/AI_tools/uml_designer/testHub/testCases.xlsx'
wb.save(outpath)

# Count cells
total = sum(1 for ws in wb.worksheets for _ in ws.iter_rows())
print(f'\nSaved: {outpath}')
print(f'Sheets: {len(wb.sheetnames)} ({", ".join(wb.sheetnames)})')
print(f'Total cells: {total}')
