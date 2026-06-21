"""TestHub API – load Excel test cases, save changes, generate test code, log reviews."""

import os
import json
from datetime import datetime
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

from app.core.config import get_settings
from app.services.llm_service import chat
from app.services.tools import clean_llm_json_response

try:
    import openpyxl
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/api/testhub", tags=["testhub"])

settings = get_settings()
TESTHUB_DIR = os.path.abspath(os.path.join(settings.uml_dir, "..", "..", "testHub"))


def _get_testhub_dir():
    d = TESTHUB_DIR
    os.makedirs(d, exist_ok=True)
    return d


# ── List Excel files ───────────────────────────────────

@router.get("/list")
async def list_test_files():
    """List all Excel files in testHub directory."""
    d = _get_testhub_dir()
    files = []
    for name in sorted(os.listdir(d)):
        if name.endswith(('.xlsx', '.xls')) and not name.startswith('~'):
            fp = os.path.join(d, name)
            files.append({
                "name": name,
                "path": fp.replace("\\", "/"),
                "size": os.path.getsize(fp),
                "modified": datetime.fromtimestamp(os.path.getmtime(fp)).isoformat(),
            })
    return {"files": files, "testhub_dir": d.replace("\\", "/")}


# ── Load Excel sheets ──────────────────────────────────

@router.get("/load")
async def load_test_file(filename: str = ""):
    """Load an Excel file and return all sheets with data."""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    d = _get_testhub_dir()
    fp = os.path.join(d, filename)
    if not os.path.exists(fp):
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")

    wb = openpyxl.load_workbook(fp)
    sheets = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c else "" for c in row]
            else:
                row_data = {}
                for j, val in enumerate(row):
                    key = headers[j] if j < len(headers) else f"col{j}"
                    row_data[key] = str(val) if val is not None else ""
                row_data["_row"] = i
                rows.append(row_data)
        sheets[sname] = {"headers": headers, "rows": rows}
    wb.close()

    # Also load relations/design info for context
    return {
        "filename": filename,
        "sheets": sheets,
        "sheet_names": list(sheets.keys()),
        "filepath": fp.replace("\\", "/"),
    }


# ── Save modified cases ────────────────────────────────

class SaveCasesRequest(BaseModel):
    filename: str
    sheets: dict  # {sheet_name: {headers: [...], rows: [...]}}


@router.post("/save")
async def save_test_file(req: SaveCasesRequest):
    """Save modified test cases back to Excel file."""
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    d = _get_testhub_dir()
    fp = os.path.join(d, req.filename)

    wb = openpyxl.load_workbook(fp)
    for sname, data in req.sheets.items():
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        headers = data.get("headers", [])
        rows = data.get("rows", [])

        # Rewrite sheet
        ws.delete_rows(1, ws.max_row)
        for j, h in enumerate(headers, 1):
            ws.cell(row=1, column=j, value=h)
        for i, row in enumerate(rows, 2):
            for j, h in enumerate(headers):
                ws.cell(row=i, column=j + 1, value=row.get(h, ""))

    wb.save(fp)
    wb.close()
    return {"success": True, "filename": req.filename}


# ── Generate test code ─────────────────────────────────

class TestGenRequest(BaseModel):
    filename: str = ""
    sheets: dict = {}
    language: str = "python"
    mode: str = "full"  # "full" or "incremental"
    changed_cases: list[dict] = []  # for incremental: [{sheet, row, changes}]


@router.post("/generate-tests")
async def generate_test_code(req: TestGenRequest):
    """Generate test code based on test cases (full or incremental)."""
    # Build a summary of test cases
    def _find_col(headers: list, keywords: list[str]) -> int:
        """Find column index by header name (case-insensitive partial match)."""
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(kw.lower() in hl for kw in keywords):
                return i
        return -1

    case_summaries = []
    for sname, data in req.sheets.items():
        headers = data.get("headers", [])
        if not headers:
            continue
        rows = data.get("rows", [])
        if not rows:
            continue

        # Use column name matching instead of hardcoded indices
        id_col = _find_col(headers, ['用例id', 'caseid', 'case_id', 'id'])
        name_col = _find_col(headers, ['用例名称', '名称', 'name', 'title'])
        priority_col = _find_col(headers, ['优先级', 'priority'])
        steps_col = _find_col(headers, ['测试步骤', '步骤', 'steps', '操作'])
        expected_col = _find_col(headers, ['预期结果', '预期', 'expected', '期望'])

        # Skip summary/overview sheets that don't contain case IDs
        if id_col < 0 or name_col < 0:
            continue

        for row in rows:
            case_id = row.get(headers[id_col], "")
            case_name = row.get(headers[name_col], "")
            priority = row.get(headers[priority_col], "") if priority_col >= 0 else ""
            steps = row.get(headers[steps_col], "") if steps_col >= 0 else ""
            expected = row.get(headers[expected_col], "") if expected_col >= 0 else ""
            case_summaries.append(
                f"[{case_id}] {case_name} (Priority: {priority})\n  Steps: {steps}\n  Expected: {expected}"
            )

    cases_text = "\n\n".join(case_summaries)  # All cases

    if req.mode == "incremental" and req.changed_cases:
        changed_text = json.dumps(req.changed_cases, ensure_ascii=False, indent=2)
        prompt = f"""Generate incremental {req.language} test code ONLY for the following CHANGED test cases:

## Changed Cases:
{changed_text}

## All Cases (for context):
{cases_text[:6000]}

Output as JSON mapping test filenames to content. Only the changed test files.
"""
    else:
        prompt = f"""Generate comprehensive {req.language} unit test code from the following test cases:

## Test Cases:
{cases_text[:8000]}

## Requirements:
- Use the standard test framework for {req.language}
- Cover all P0 and P1 test cases
- Return as JSON mapping filenames to content
- Only output the JSON object.
"""

    response = await chat(
        prompt=prompt,
        system_prompt=f"You are an expert {req.language} test engineer. Output only valid JSON.",
        temperature=0.3,
        max_tokens=8192,
    )

    try:
        cleaned = clean_llm_json_response(response)
        result = json.loads(cleaned)
        return {"files": result, "language": req.language, "mode": req.mode}
    except json.JSONDecodeError:
        return {"files": {"test_generated.py": response}, "language": req.language, "mode": req.mode}


# ── Save test review log ───────────────────────────────

class TestReviewRequest(BaseModel):
    action: str = ""  # "view", "edit", "generate_full", "generate_incremental", "accept", "reject"
    comment: str = ""
    filename: str = ""
    sheet: str = ""
    case_id: str = ""
    details: str = ""


@router.post("/save-review")
async def save_test_review(req: TestReviewRequest):
    """Save test case review operation log to test_review.txt."""
    d = _get_testhub_dir()
    review_file = os.path.join(d, "test_review.txt")

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = f"[{ts}] {req.action}"
    if req.filename:
        entry += f" | File: {req.filename}"
    if req.sheet:
        entry += f" | Sheet: {req.sheet}"
    if req.case_id:
        entry += f" | Case: {req.case_id}"
    if req.comment:
        entry += f"\n  Comment: {req.comment}"
    if req.details:
        entry += f"\n  Details: {req.details}"
    entry += "\n"

    with open(review_file, "a", encoding="utf-8") as f:
        f.write(entry)

    return {"success": True, "file": review_file.replace("\\", "/")}
